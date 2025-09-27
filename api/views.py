import csv

from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook
from rest_framework_simplejwt.tokens import RefreshToken

from api.serializer import UploadFileSerializer, ASMRegisterSerializer, LoginSerializer
from asm.permissions import IsAdminAPI
from master.models import State, District, Office, PincodeData


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions

from zonemanager.serializer import ZoneManagerSerializer
from .models import UserProfile
from .serializer import UserProfileSerializer
from zonemanager.models import ZoneManager

class UploadAndProcessFileAPI(APIView):
    def post(self, request, *args, **kwargs):
        serializer = UploadFileSerializer(data=request.data)
        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            expected_headers = [
                "circlename", "regionname", "divisionname", "officename", "pincode",
                "officetype", "delivery", "district", "statename", "latitude", "longitude"
            ]

            try:
                rows = []

                # --- CSV ---
                if uploaded_file.name.endswith('.csv'):
                    decoded_file = uploaded_file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)

                    # Normalize headers
                    headers = [h.lower() for h in reader.fieldnames]
                    if not all(h in headers for h in expected_headers):
                        return Response({"error": "CSV missing required headers"},
                                        status=status.HTTP_400_BAD_REQUEST)

                    rows = reader

                # --- XLSX ---
                elif uploaded_file.name.endswith('.xlsx'):
                    workbook = load_workbook(uploaded_file)
                    sheet = workbook.active
                    header = [str(cell.value).strip().lower() for cell in sheet[1]]

                    if not all(h in header for h in expected_headers):
                        return Response({"error": "XLSX missing required headers"},
                                        status=status.HTTP_400_BAD_REQUEST)

                    rows = (
                        dict(zip(header, [
                            sheet.cell(row=r, column=c).value
                            for c in range(1, sheet.max_column + 1)
                        ]))
                        for r in range(2, sheet.max_row + 1)
                    )

                else:
                    return Response({"error": "Unsupported file type"},
                                    status=status.HTTP_400_BAD_REQUEST)

                # --- Process Rows ---
                created_count, updated_count, skipped_count = 0, 0, 0

                for row in rows:
                    pincode = str(row.get("pincode")).strip() if row.get("pincode") else None
                    if not pincode:
                        skipped_count += 1
                        continue

                    # Handle NA values for latitude/longitude
                    def clean_float(value):
                        if value is None:
                            return None
                        try:
                            v = str(value).strip().upper()
                            if v in ("NA", "", "NONE", "NULL"):
                                return None
                            return float(v)
                        except ValueError:
                            return None

                    latitude = clean_float(row.get("latitude"))
                    longitude = clean_float(row.get("longitude"))

                    # Create related objects
                    state, _ = State.objects.get_or_create(name=row["statename"])
                    district, _ = District.objects.get_or_create(name=row["district"], state=state)

                    # Use update_or_create instead of get_or_create for Office
                    office, _ = Office.objects.update_or_create(
                        name=row["officename"],
                        district=district,
                        defaults={
                            "pincode": pincode,
                            "officetype": row["officetype"],
                        }
                    )

                    # Handle NA / None values for latitude & longitude
                    latitude = None if str(row.get("latitude")).strip().lower() in ["na", "", "none"] else float(
                        row["latitude"])
                    longitude = None if str(row.get("longitude")).strip().lower() in ["na", "", "none"] else float(
                        row["longitude"])

                    # Use update_or_create for PincodeData as well
                    obj, created = PincodeData.objects.update_or_create(
                        pincode=pincode,
                        defaults={
                            "officename": row["officename"],
                            "district": row["district"],
                            "statename": row["statename"],
                            "officetype": row["officetype"],
                            "delivery": row["delivery"],
                            "latitude": latitude,
                            "longitude": longitude,
                        }
                    )

                    if created:
                        created_count += 1
                    else:
                        skipped_count += 1  # actually updated, not skipped

                return Response({
                    "message": "File processed successfully",
                    "created": created_count,
                    "updated": updated_count,
                    "skipped": skipped_count
                }, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ASMRegisterAPI(APIView):
    def post(self, request, *args, **kwargs):
        serializer = ASMRegisterSerializer(data=request.data)
        if serializer.is_valid():
            asm = serializer.save()
            return Response(
                {"message": "ASM registered successfully", "asm_id": asm.id},
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class LoginView(APIView):
    permission_classes = [AllowAny]
    def post(self, request, *args, **kwargs):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data["user"]
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)
            role = ""
            if user.groups.filter(name__iexact='admin').exists():
                role = 'Admin'
            elif user.groups.filter(name__iexact='Area Sales Manager').exists():
                role = 'Area Sales Manager'
            elif user.groups.filter(name__iexact='Zonal Manager').exists():
                role = 'Zonal Manager'
            elif user.groups.filter(name__iexact='Technical Manager').exists():
                role = 'Technical Manager'
            else:
                role = 'Unable to find user role'
            response=Response({
                "message": "Login successful",
                "token": access_token,
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "name": user.first_name + " " + user.last_name,
                    "email": user.email,
                    "is_active": user.is_active,
                    "role":role
                }
            }, status=status.HTTP_200_OK)

            response.set_cookie(
                key="refresh_token",
                value=str(refresh),
                httponly=True,  # JS cannot access
                secure=False,  # set True in production with HTTPS
                samesite="Lax",  # or "Strict" / "None" depending on frontend
                max_age=7 * 24 * 60 * 60  # 7 days
            )

            # ✅ Send access token in header
            response["Authorization"] = f"Bearer {access_token}"
            return response

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class RefreshTokenView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        refresh_token = request.COOKIES.get("refresh_token")
        if refresh_token:
            try:
                # Verify the refresh token
                token = RefreshToken(refresh_token)
                # Generate a new access token
                access_token = str(token.access_token)
                return Response({"access_token": access_token}, status=200)
            except Exception as e:
                return Response({"error": "Invalid refresh token"}, status=400)
        return Response({"error": "Refresh token not found"}, status=400)


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        response = Response({"success":True, "message": "Logout successful"}, status=status.HTTP_200_OK)
        response.delete_cookie("refresh_token")
        return response





class UserProfileAPI(APIView):
    """
    API to get the UserProfile of the currently logged-in user
    """
    permission_classes = [permissions.IsAuthenticated]  # Only logged-in users can access

    def get(self, request):
        try:
            profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
            return Response({"detail": "UserProfile not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = UserProfileSerializer(profile)
        response_data = {}
        if request.user.groups.filter(name__iexact='Zonal Manager').exists():
            zone_manager = ZoneManager.objects.get(user=request.user)
            if zone_manager:
                zone_serializer = ZoneManagerSerializer(zone_manager)
                response_data = {
                    "role_id": zone_serializer.data["id"],
                    "states": zone_serializer.data["states"],
                    "offices": zone_serializer.data["offices"],
                    "districts": zone_serializer.data["districts"],
                    "user_id": serializer.data["user"]["id"],
                    "username": serializer.data["user"]["username"],
                    "name": serializer.data["user"]["full_name"],
                    "email": serializer.data["user"]["email"],
                    "role": serializer.data["user"]["groups"][0],
                    "employee_id": serializer.data["employee_id"],
                    "phone": serializer.data["phone"],
                    "address": serializer.data["address"],
                    "date_of_birth": serializer.data["date_of_birth"],
                    "blood_group": serializer.data["blood_group"],
                    "join_date": serializer.data["join_date"],
                    "avatar": serializer.data["avatar"],
                    "department": serializer.data["department"],
                    "designation": serializer.data["designation"],
                    "employee_status": serializer.data["employee_status"]
                }


        if response_data=={}:
            return Response({"success":False, "message": "UserRole not found"}, status=status.HTTP_404_NOT_FOUND)

        return Response(
            {
                "success":True,
                "message": "User profile fetched successfully",
                "data":response_data
            }
            , status=status.HTTP_200_OK)

